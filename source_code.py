from tkinter import *
import os
import numpy as np
import time
from tqdm.auto import tqdm            
from openpyxl.workbook import *
from openpyxl import load_workbook




def fn1():
    def fn2():
        root=Tk()
        root.geometry("300x100")
        la1=Label(root,text="successfully created",font=150).grid(row=0,column=1)
        root.mainloop()
        
    fileObj = open("a.txt", "r") #opens the file in read mode

    words = fileObj.read().splitlines() #puts the file into an array


    fileObj.close()
    first=[]
    for i in range(0,len(words)):
        if (words[i][5]+words[i][6]=="02" and words[i][8]+words[i][9]=="31") or (words[i][5]+words[i][6]=="04" and words[i][8]+words[i][9]=="31") or (words[i][5]+words[i][6]=="06" and words[i][8]+words[i][9]=="31") or (words[i][5]+words[i][6]=="09" and words[i][8]+words[i][9]=="31") or (words[i][5]+words[i][6]=="11" and words[i][8]+words[i][9]=="31"):
            pass
        else:
            first.append(words[i])

    second=[]
    for i in range(0,len(first)):
        if (first[i][5]+first[i][6]=="02" and first[i][8]+first[i][9]=="30"):
            pass
        else:
            second.append(first[i])


    final=[]

    for i in range(0,len(second)):
        if int(second[i][0]+second[i][1]+second[i][2]+second[i][3])%4!=0 and second[i][5]+second[i][6]=="02" and second[i][8]+second[i][9]=="29":
            pass
            
        
        else:
            final.append(second[i])


    fopen=open("%s.txt" %(name_file),"a")

    for i in final:
        fopen.write(i+"\n")
        
    fopen.close()
    fn2()
    

############################################################
try:
    os.remove("a.txt")
except:
    pass

textfile=open("a.txt","a")
    
#########1





wb1=load_workbook('max.xlsx' )
max_sheet=wb1['Sheet']
wb3=load_workbook('rain.xlsx' )
rain_sheet=wb3['Sheet']
wb2=load_workbook('min.xlsx' )
min_sheet=wb2['Sheet']

name_file=input("Name of station = ")
############additional
print("-----------------------------")
for i in range(8,len(max_sheet['A'])):
    if max_sheet.cell(row=i,column=3).value==name_file:
        very_1_value=max_sheet.cell(row=i+1,column=2).value
        very_1_row=i+1
        break
for i in range(very_1_row,len(max_sheet['A'])):
    if max_sheet.cell(row=i,column=2).value==":" or max_sheet.cell(row=i,column=1).value=="Station":
        very_2_row=i-1
        break


for i in range(very_1_row,very_2_row-12,12):
    tr1=max_sheet.cell(row=i,column=2).value
    tr2=max_sheet.cell(row=i+12,column=2).value 
    if tr2!=tr1+1:
        print(f'daily maximum temperature data misses from {tr1} to {tr2}')
print("-----------------------------")        
print("-----------------------------")
for i in range(6,len(min_sheet['A'])):
    if min_sheet.cell(row=i,column=3).value==name_file:
        very_1_value2=min_sheet.cell(row=i+3,column=2).value
        very_1_row2=i+3
        break
for i in range(very_1_row2,len(min_sheet['A'])):
    if min_sheet.cell(row=i,column=2).value==":" or min_sheet.cell(row=i,column=1).value=="Station":
        very_2_row2=i-1
        break


for i in range(very_1_row2,very_2_row2-12,12):
    tr11=min_sheet.cell(row=i,column=2).value
    tr22=min_sheet.cell(row=i+12,column=2).value 
    if tr22!=tr11+1:
        print(f'daily minimum temperature data misses from {tr1} to {tr2}')
print("-----------------------------") 
print("-----------------------------")
for i in range(9,len(rain_sheet['A'])):
    if rain_sheet.cell(row=i,column=1).value==name_file:
        very_1_value3=rain_sheet.cell(row=i,column=2).value
        very_1_row3=i
        break
for i in range(very_1_row3,len(rain_sheet['A'])):
    if rain_sheet.cell(row=i,column=1).value!=name_file:
        very_2_row3=i
        break


for i in range(very_1_row3,very_2_row3-12,12):
    tr13=rain_sheet.cell(row=i,column=2).value
    tr23=rain_sheet.cell(row=i+12,column=2).value 
    if tr23!=tr13+1:
        print(f'daily rainfall data misses from {tr1} to {tr2}')
print("-----------------------------")    
       
     
'''
print(very_1_row)
print(very_2_row)
'''

start_p=int(input("base period start from = "))   
end_p=int(input("base period end at = ")) 
base_p=end_p-start_p+1 






#######--------
for i in range(8,len(max_sheet['A'])):
    if max_sheet.cell(row=i,column=3).value==name_file:
        
        initial_index_max=i+1  
        break
for i in range(initial_index_max,len(max_sheet['A'])):
    if max_sheet.cell(row=i,column=2).value==start_p and max_sheet.cell(row=i,column=3).value==1:
        lim1_max=i
        break

lim2_max=lim1_max+base_p*12-1
        
#######-----

#######--------
for i in range(6,len(min_sheet['A'])):
    if min_sheet.cell(row=i,column=3).value==name_file:
        
        initial_index_min=i+3  
        break
for i in range(initial_index_min,len(min_sheet['A'])):
    if min_sheet.cell(row=i,column=2).value==start_p and min_sheet.cell(row=i,column=3).value==1:
        lim1_min=i
        break

lim2_min=lim1_min+base_p*12-1
   
#######-----


clim=int()
#######--------
for i in range(8,len(rain_sheet['A'])):
    if rain_sheet.cell(row=i,column=1).value==name_file and rain_sheet.cell(row=i,column=2).value==start_p and rain_sheet.cell(row=i,column=3).value==1  :
        clim=i
lim1_rain=clim      
lim2_rain=lim1_rain+base_p*12-1        
#######-----
if max_sheet.cell(row=lim2_max,column=2).value!=end_p :
    root=Tk()
    root.geometry("300x100")
    la1=Label(root,text="Daily Maximum temp data missing check the database",font=150).grid(row=0,column=1)
    text2="%s :" %(name_file) 
    text3="%d to " %(start_p) 
    text4="%d" %(end_p)
    text=text2+text3+text4
    max_sheet['A1'].hyperlink = "#Sheet!B%d" %(lim1_max)
    max_sheet['A1'].value=text
    wb1.save(r"max.xlsx")
    root.mainloop()
    
elif  min_sheet.cell(row=lim2_min,column=2).value!=end_p:
    root=Tk()
    root.geometry("300x100")
    la1=Label(root,text="Daily Minimum temp data missing check the database",font=150).grid(row=0,column=1)
    text2="%s :" %(name_file) 
    text3="%d to " %(start_p) 
    text4="%d" %(end_p)
    text=text2+text3+text4
    min_sheet['A1'].hyperlink = "#Sheet!B%d" %(lim1_min)
    min_sheet['A1'].value=text
    wb2.save(r"min.xlsx")
    root.mainloop()



elif rain_sheet.cell(row=lim2_rain,column=2).value!=end_p: 
    root=Tk()
    root.geometry("300x100")
    la1=Label(root,text="Daily Rainfall data missing check the database",font=150).grid(row=0,column=1)
    text2="%s :" %(name_file) 
    text3="%d to " %(start_p) 
    text4="%d" %(end_p)
    text=text2+text3+text4
    rain_sheet['A1'].hyperlink = "#Sheet!B%d" %(lim1_rain)
    rain_sheet['A1'].value=text
    wb3.save(r"rain.xlsx")
    root.mainloop()
else:
    maxtemp=[]
    for i in range(lim1_max,lim2_max+1):
        for j in range(4,35):
            if max_sheet.cell(row=i,column=j).value==None or max_sheet.cell(row=i,column=j).value=="****":
                maxtemp.append(-99.9)
            else:
                maxtemp.append(float(max_sheet.cell(row=i,column=j).value))
                
    print(len(maxtemp))            
                
            
    ###########2







    mintemp=[]
    for i in range(lim1_min,lim2_min+1):
        for j in range(4,35):
            if min_sheet.cell(row=i,column=j).value==None or min_sheet.cell(row=i,column=j).value=="****":
                mintemp.append(-99.9)
            else:
                mintemp.append(float(min_sheet.cell(row=i,column=j).value))
        
    ############3








    rain=[]
    for i in range(lim1_rain,lim2_rain+1):
        for j in range(4,35):
            if rain_sheet.cell(row=i,column=j).value==None or rain_sheet.cell(row=i,column=j).value=="***":
                rain.append(-99.9)
            else:
                rain.append(float(rain_sheet.cell(row=i,column=j).value))
       
    #####
    xx=(end_p-start_p)
    y=[]
    for i in range(0,xx+1):
        for j in range(0,372):
            y.append(start_p)
        start_p=start_p+1
        
    days=[]
    for i in range(0,12*(xx+1)):
        
        for j in range(1,32):
            days.append(format(j,'02d'))

    month=[]
    k=1    
    for i in range(0,12*(xx+1)):
        
        for j in range(1,32):
            month.append(format(k,'02d'))
        if k==12:
            k=1
        else:
            k=k+1


    for i in range(0,len(maxtemp)):
        textfile.write(str(y[i])+" "+str(month[i])+" "+str(days[i])+" "+str(rain[i])+" "+str(maxtemp[i])+" "+str(mintemp[i])+" "+"\n")
        
    textfile.close() 

        

    
    fn1()


    ##############--------



    


